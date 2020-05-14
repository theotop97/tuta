import os

from openpyxl import load_workbook

class ParseExcel:

    def __init__(self, excelName, sheetName):
        excelPath = os.path.abspath(r"../MyAppiumCode/excel") + r"/" + excelName
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRow = self.sheet.max_row

    def getDataFromSheet(self):
        dataList = []
        for line in list(self.sheet.rows)[1:]:
            tempList = [line[1].value, line[2].value]
            dataList.append(tempList)
        return dataList

if __name__ == '__main__':
    pe = ParseExcel("settingsearch.xlsx", "search1")
    for i in pe.getDataFromSheet():
        print(i[0], i[1])
